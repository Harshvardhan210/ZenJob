import os
import uuid
import shutil
import json
import time
import threading
import hashlib
from typing import Optional, List
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables from backend/.env
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, ".env"))


from fastapi import FastAPI, UploadFile, File, Header, HTTPException, status, Depends, Response, Request
from .auth import get_current_user
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from jose import JWTError, jwt

import pandas as pd
import io
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from . import database, extraction, models

def get_gemini_key():
    """Retrieves the Gemini API key from environment variables."""
    return os.getenv("GEMINI_API_KEY")


# 0. In-Process TTL Cache
class SimpleCache:
    """Thread-safe in-memory TTL cache."""
    def __init__(self, ttl: int = 30):
        self._store: dict = {}
        self._lock = threading.Lock()
        self.ttl = ttl

    def get(self, key: str):
        with self._lock:
            entry = self._store.get(key)
            if entry and (time.time() - entry["ts"]) < self.ttl:
                return entry["val"]
            return None

    def set(self, key: str, value):
        with self._lock:
            self._store[key] = {"val": value, "ts": time.time()}

    def delete(self, key: str):
        with self._lock:
            self._store.pop(key, None)

    def delete_prefix(self, prefix: str):
        with self._lock:
            keys = [k for k in self._store if k.startswith(prefix)]
            for k in keys:
                del self._store[k]

# jobs_cache: caches GET /api/jobs and GET /api/resumes per user (30s TTL)
jobs_cache = SimpleCache(ttl=30)
# extract_cache: caches AI extraction results by content hash (5min TTL)
extract_cache = SimpleCache(ttl=300)

# 1. Initialize FastAPI App
app = FastAPI(
    title="ZenJob API",
    description="Backend API for extracting and collecting job application posters/screenshots",
    version="1.0.0"
)

# 2. Configure CORS
allowed_origins_env = os.getenv("ALLOWED_ORIGINS")
if allowed_origins_env:
    origins = [o.strip() for o in allowed_origins_env.split(",") if o.strip()]
else:
    origins = [
        "http://localhost:5173",
        "http://localhost:5000",
        "http://localhost:3000",
        "http://localhost:8000",
        "capacitor://localhost",
        "http://localhost",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 2.5. HTTP Security Headers Middleware
@app.middleware("http")
async def add_security_headers(request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Content-Security-Policy"] = "default-src 'self' 'unsafe-inline' 'unsafe-eval' https://fonts.googleapis.com https://fonts.gstatic.com; img-src 'self' data: blob: http: https:;"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response

# 3. Directories Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
RESUMES_DIR = os.path.join(UPLOAD_DIR, "resumes")
os.makedirs(RESUMES_DIR, exist_ok=True)

def verify_path_safety(file_path: str, base_directory: str = UPLOAD_DIR) -> bool:
    """
    Checks that the given file path is strictly located within the base_directory
    to prevent directory traversal attacks.
    """
    try:
        resolved_path = os.path.abspath(file_path)
        resolved_base = os.path.abspath(base_directory)
        return os.path.commonpath([resolved_base, resolved_path]) == resolved_base
    except Exception:
        return False

# Mount the uploads directory to serve images statically to the React frontend
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# 4.5. Session Security Settings
# Endpoints for Gemini API key management have been removed.
# Local extraction logic is used instead.

def inject_resume_match(extracted_data: dict, user_id: str) -> dict:
    """
    Checks the database for an active resume. If one exists, calls extraction.match_resume_to_jd
    and injects match metrics into the extraction output dictionary.
    """
    try:
        resumes = database.get_all_resumes(user_id)
        active_resume = None
        for r in resumes:
            if r.get("is_active") == 1 or r.get("is_active") == True:
                active_resume = r
                break
                
        if active_resume:
            filename = active_resume.get("filepath").split("/")[-1]
            resume_path = os.path.join(RESUMES_DIR, filename)
            
            match_res = extraction.match_resume_to_jd(
                resume_path=resume_path,
                jd_details=extracted_data
            )
            
            if match_res:
                extracted_data["match_score"] = match_res.get("match_score")
                extracted_data["matching_skills"] = match_res.get("matching_skills", [])
                extracted_data["missing_skills"] = match_res.get("missing_skills", [])
                extracted_data["suggestions"] = match_res.get("suggestions", [])
    except Exception as e:
        print(f"Error during real-time resume matching: {str(e)}")
        
    return extracted_data

@app.post("/api/extract", response_model=models.JobExtraction)
async def extract_job(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user)
):
    """
    Uploads a job poster image/screenshot, saves it locally,
    and runs local OCR and NLP to extract structured job details.
    """
    # Verify file is an image
    content_type = file.content_type or ""
    if not content_type.startswith("image/"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file must be an image (PNG, JPG, JPEG, WEBP, etc.)"
        )
        
    # Verify file size limit (10MB) to prevent Denial of Service (DoS)
    try:
        file.file.seek(0, 2)
        file_size = file.file.tell()
        file.file.seek(0)
        if file_size > 10 * 1024 * 1024:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded image flyer size exceeds the 10MB limit."
            )
    except HTTPException as he:
        raise he
    except Exception:
        # Fallback if seek/tell fails on certain file wrappers
        pass
        
    # Generate unique filename to avoid overwrites
    file_ext = os.path.splitext(file.filename)[1] or ".png"
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    saved_image_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    # Save file locally
    try:
        with open(saved_image_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save uploaded image: {str(e)}"
        )
        
    # Perform extraction
    try:
        extracted_data = extraction.extract_job_info_from_image(
            image_path=saved_image_path
        )
        
        # Attach the relative image path in the response headers or custom body wrapper 
        # so frontend can read it. Let's return the extracted data, but we can pass 
        # the saved image path as an additional field or custom header. 
        # To make it easy, we'll append the image name to additional notes or we can return 
        # a wrapper object. Let's return the extracted data directly, but we can add a response header 
        # or handle saving in frontend. Wait, let's include the saved image path as part of the return!
        # Since JobExtraction doesn't have image_path by default, let's convert to dict, insert it, 
        # and return it. Wait, the frontend can just send the image_path back when saving the job.
        
        # Let's return a dictionary that matches JobExtraction but also has the "image_path"!
        response_dict = extracted_data.model_dump()
        response_dict["image_path"] = f"/uploads/{unique_filename}"
        response_dict = inject_resume_match(response_dict, user_id=user_id)
        return response_dict
        
    except ValueError as ve:
        # Clean up file if extraction failed due to missing API key / invalid image
        if os.path.exists(saved_image_path):
            try:
                os.remove(saved_image_path)
            except Exception:
                pass
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(ve)
        )
    except Exception as e:
        if os.path.exists(saved_image_path):
            try:
                os.remove(saved_image_path)
            except Exception:
                pass
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI Extraction failed: {str(e)}"
        )

@app.post("/api/extract-text", response_model=models.JobExtraction)
def extract_job_text(
    request_data: Request,
    request: models.TextExtractionRequest,
    user_id: str = Depends(get_current_user)
):
    """
    Accepts raw job description text and runs local NLP to extract structured job details.
    Results are cached by content hash for 5 minutes.
    """
    if not request.text.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Job description text cannot be empty"
        )

    # Cache key: hash of text + user so different users stay isolated
    cache_key = "text:" + hashlib.sha256(f"{user_id}:{request.text}".encode()).hexdigest()
    cached = extract_cache.get(cache_key)
    if cached is not None:
        return cached

    try:
        extracted_data = extraction.extract_job_info_from_text(
            text=request.text
        )
        response_dict = extracted_data.model_dump()
        response_dict = inject_resume_match(response_dict, user_id=user_id)
        extract_cache.set(cache_key, response_dict)
        return response_dict
    except ValueError as ve:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(ve)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Text AI Extraction failed: {str(e)}"
        )

@app.post("/api/extract-url", response_model=models.JobExtraction)
def extract_job_url(
    request_data: Request,
    request: models.UrlExtractionRequest,
    user_id: str = Depends(get_current_user)
):
    """
    Accepts a URL, scrapes the webpage for text content, and runs local NLP to extract structured job details.
    Results are cached by URL hash for 5 minutes.
    """
    if not request.url.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="URL cannot be empty"
        )

    # Cache key: hash of URL + user
    cache_key = "url:" + hashlib.sha256(f"{user_id}:{request.url.strip()}".encode()).hexdigest()
    cached = extract_cache.get(cache_key)
    if cached is not None:
        return cached
        
    try:
        # Fetch the webpage
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        response = requests.get(request.url, headers=headers, timeout=15)
        response.raise_for_status()
        
        # Parse HTML and extract clean text
        soup = BeautifulSoup(response.content, "html.parser")
        
        # Remove script and style elements
        for script in soup(["script", "style", "noscript", "header", "footer", "nav"]):
            script.extract()
            
        text = soup.get_text(separator=' ')
        
        # Collapse whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        clean_text = '\\n'.join(chunk for chunk in chunks if chunk)
        
        if len(clean_text) < 50:
            raise ValueError("Could not extract enough meaningful text from this URL. It might be heavily javascript-rendered.")
            
        # Run AI extraction
        extracted_data = extraction.extract_job_info_from_text(
            text=clean_text
        )
        # Always populate the application_link with the source URL if it wasn't found
        if not extracted_data.application_link:
            extracted_data.application_link = request.url
            
        response_dict = extracted_data.model_dump()
        response_dict = inject_resume_match(response_dict, user_id=user_id)
        extract_cache.set(cache_key, response_dict)
        return response_dict
        
    except requests.exceptions.RequestException as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to fetch URL: {str(e)}"
        )
    except ValueError as ve:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(ve)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"URL AI Extraction failed: {str(e)}"
        )

@app.get("/api/jobs", response_model=List[models.JobDB])
def get_jobs(user_id: str = Depends(get_current_user)):
    """Retrieves all collected job applications. Cached per user for 30 seconds."""
    cache_key = f"jobs:{user_id}"
    cached = jobs_cache.get(cache_key)
    if cached is not None:
        return cached
    try:
        result = database.get_all_jobs(user_id)
        jobs_cache.set(cache_key, result)
        return result
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve jobs: {str(e)}"
        )

@app.post("/api/jobs", response_model=models.JobDB)
def create_job(job: models.JobExtraction, image_path: Optional[str] = None, user_id: str = Depends(get_current_user)):
    """Saves a confirmed job application to the database."""
    try:
        job_dict = job.model_dump()
        job_dict["image_path"] = image_path
        job_dict["user_id"] = user_id
        
        job_id = database.add_job(job_dict, user_id)
        saved_job = database.get_job_by_id(job_id, user_id)
        if not saved_job:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Saved job could not be retrieved"
            )
        jobs_cache.delete(f"jobs:{user_id}")  # Invalidate jobs list cache
        return saved_job
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save job: {str(e)}"
        )

@app.put("/api/jobs/{job_id}", response_model=models.JobDB)
def update_job(job_id: str, job: models.JobExtraction, image_path: Optional[str] = None, user_id: str = Depends(get_current_user)):
    """Updates an existing job application in the database."""
    try:
        existing = database.get_job_by_id(job_id, user_id)
        if not existing:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Job with ID {job_id} not found"
            )
            
        job_dict = job.model_dump()
        # Keep previous image path if not provided
        job_dict["image_path"] = image_path if image_path is not None else existing.get("image_path")
        job_dict["user_id"] = user_id
        
        success = database.update_job(job_id, job_dict, user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to update job in database"
            )
            
        jobs_cache.delete(f"jobs:{user_id}")  # Invalidate jobs list cache
        updated = database.get_job_by_id(job_id, user_id)
        return updated
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update job: {str(e)}"
        )

@app.delete("/api/jobs/{job_id}")
def delete_job(job_id: str, user_id: str = Depends(get_current_user)):
    """Deletes a job application and its associated image file."""
    try:
        job = database.get_job_by_id(job_id, user_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Job with ID {job_id} not found"
            )
            
        # Try to delete original image if it exists
        img_rel_path = job.get("image_path")
        if img_rel_path:
            # Strip leading slash if present
            if img_rel_path.startswith("/"):
                img_rel_path = img_rel_path[1:]
            full_img_path = os.path.abspath(os.path.join(BASE_DIR, img_rel_path))
            if verify_path_safety(full_img_path, UPLOAD_DIR):
                if os.path.exists(full_img_path):
                    try:
                        os.remove(full_img_path)
                    except Exception:
                        pass  # Fail silently if image deletion fails, focus on database delete
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid or unauthorized image file path"
                )
                    
        success = database.delete_job(job_id, user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to delete job from database"
            )
        jobs_cache.delete(f"jobs:{user_id}")  # Invalidate jobs list cache
        return {"status": "success", "message": f"Job {job_id} deleted successfully"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete job: {str(e)}"
        )

@app.get("/api/export")
def export_jobs_excel(user_id: str = Depends(get_current_user)):
    """
    Fetches all jobs in the SQLite database, compiles them into a beautiful, 
    highly styled, formatted Excel spreadsheet using Pandas & OpenPyXL, and 
    returns it as a streaming download.
    """
    try:
        jobs = database.get_all_jobs(user_id)
        if not jobs:
            # Return a simple empty excel instead of error, or prompt user
            df = pd.DataFrame(columns=[
                "Company Name", "Job Role", "Email", "Phone", "Location", 
                "Job Type", "Work Mode", "Skills", "Experience Required", 
                "Application Link", "Notes", "Status", "Extracted At"
            ])
        else:
            # Format skills list into a comma-separated string for excel clarity
            formatted_jobs = []
            for j in jobs:
                skills_val = j.get("skills", [])
                skills_str = ", ".join(skills_val) if isinstance(skills_val, list) else str(skills_val)
                
                # Format date nicely
                raw_date = j.get("extracted_at", "")
                try:
                    dt = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S")
                    date_str = dt.strftime("%d-%b-%Y %I:%M %p")
                except Exception:
                    date_str = raw_date
                    
                formatted_jobs.append({
                    "Company Name": j.get("company_name") or "",
                    "Job Role": j.get("job_role") or "",
                    "Email": j.get("email") or "",
                    "Phone": j.get("phone") or "",
                    "Location": j.get("location") or "",
                    "Job Type": j.get("job_type") or "",
                    "Work Mode": j.get("work_mode") or "",
                    "Skills": skills_str,
                    "Experience Required": j.get("experience_required") or "",
                    "Application Link": j.get("application_link") or "",
                    "Notes": j.get("additional_notes") or "",
                    "Status": j.get("application_status") or "Applied",
                    "Extracted At": date_str
                })
            df = pd.DataFrame(formatted_jobs)
            
        # Create an in-memory buffer for Excel file
        output = io.BytesIO()
        
        # Write to excel using OpenPyXL engine inside Pandas
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Collected Jobs", index=False)
            
            # Access underlying Workbook and Sheet to style it
            workbook = writer.book
            worksheet = writer.sheets["Collected Jobs"]
            
            # Design Elements: Slate Indigo theme
            header_fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")  # Deep Indigo #312e81
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            zebra_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")  # Extremely light violet
            border_thin = Side(border_style="thin", color="E5E7EB")  # Light gray
            cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            
            # Style header row (Row 1)
            worksheet.row_dimensions[1].height = 26
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = cell_border
                
            # Style data rows
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 20
                is_even = (row_idx % 2 == 0)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = Font(name="Calibri", size=10)
                    cell.border = cell_border
                    
                    # Apply light zebra striping
                    if is_even:
                        cell.fill = zebra_fill
                        
                    # Alignment defaults
                    col_name = df.columns[col_idx - 1]
                    if col_name in ["Email", "Phone", "Job Type", "Work Mode", "Extracted At"]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
            # Autofit column widths dynamically
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or "")
                    if cell.row == 1:
                        max_len = max(max_len, len(val) + 4)  # Give extra padding to headers
                    else:
                        max_len = max(max_len, len(val))
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)  # Constrained between 12 and 45 chars
                
            # Add Data Validation for the Status column (Column L)
            status_dv = DataValidation(type="list", formula1='"Applied,Test Process,Screening,Pending response,Rejected,Selected"', allow_blank=True)
            # Apply to all rows from row 2 to 1000 in column 'L'
            status_dv.add(f'L2:L1000')
            worksheet.add_data_validation(status_dv)
                
            # Add Conditional Formatting for Status colors
            green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            green_font = Font(color='065F46')
            
            red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            red_font = Font(color='991B1B')
            
            orange_fill = PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid')
            orange_font = Font(color='9A3412')
            
            purple_fill = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')
            purple_font = Font(color='6B21A8')
            
            yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            yellow_font = Font(color='92400E')
            
            indigo_fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
            indigo_font = Font(color='3730A3')

            worksheet.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Selected"'], stopIfTrue=True, fill=green_fill, font=green_font))
            worksheet.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Rejected"'], stopIfTrue=True, fill=red_fill, font=red_font))
            worksheet.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Test Process"'], stopIfTrue=True, fill=orange_fill, font=orange_font))
            worksheet.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Screening"'], stopIfTrue=True, fill=purple_fill, font=purple_font))
            worksheet.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Pending response"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
            worksheet.conditional_formatting.add('L2:L1000', CellIsRule(operator='equal', formula=['"Applied"'], stopIfTrue=True, fill=indigo_fill, font=indigo_font))
                
            # Ensure grid lines are visible
            worksheet.views.sheetView[0].showGridLines = True
            
        # Return file as a streaming download response
        output.seek(0)
        
        filename = f"ZenJob_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate Excel report: {str(e)}"
        )

@app.post("/api/jobs/{job_id}/analyze")
def analyze_job_match(
    request: Request,
    job_id: str,
    user_id: str = Depends(get_current_user)
):
    """
    Fetches an existing job, finds the user's active resume, and conducts 
    a deep AI match analysis using Gemini.
    """
    api_key = get_gemini_key()
    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED, 
            detail="Gemini match analysis is disabled (GEMINI_API_KEY not set). Using local matching instead."
        )


    # 1. Fetch Job
    job = database.get_job_by_id(job_id, user_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job with ID {job_id} not found"
        )
        
    # 2. Find Active Resume
    resumes = database.get_all_resumes(user_id)
    active_resume = None
    for r in resumes:
        if r.get("is_active") == 1 or r.get("is_active") == True:
            active_resume = r
            break
            
    if not active_resume:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No active resume found. Please upload a resume and set it as active in the Resumes section."
        )
        
    # 3. Path Management
    filename = active_resume.get("filepath").split("/")[-1]
    resume_path = os.path.join(RESUMES_DIR, filename)
    
    if not os.path.exists(resume_path):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="The active resume file could not be found on the server."
        )
        
    # 4. Conduct Matching
    try:
        match_res = extraction.match_resume_to_jd(
            resume_path=resume_path,
            jd_details=job
        )
        
        if not match_res:
            raise ValueError("AI failed to return valid analysis results.")
            
        return match_res
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Analysis failed: {str(e)}"
        )

# --- RESUME MANAGEMENT API ENDPOINTS ---

@app.post("/api/resumes", response_model=models.ResumeDB)
async def upload_resume(file: UploadFile = File(...), user_id: str = Depends(get_current_user)):
    """
    Uploads a resume file (PDF, DOC, DOCX), saves it to the local resumes directory,
    and stores its metadata in the database.
    """
    # 1. Validate file extension
    filename = file.filename or "resume.pdf"
    file_ext = os.path.splitext(filename)[1].lower()
    if file_ext not in [".pdf", ".doc", ".docx"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Only PDF, DOC, and DOCX resumes are supported."
        )

    # 2. Prevent naming conflicts by generating a unique disk filename
    unique_id = uuid.uuid4().hex[:8]
    base_name = os.path.splitext(filename)[0]
    # Clean up base name a bit (remove spaces/special characters)
    clean_base = "".join(c for c in base_name if c.isalnum() or c in ["-", "_"]).strip()
    disk_filename = f"{clean_base}_{unique_id}{file_ext}"
    
    saved_file_path = os.path.join(RESUMES_DIR, disk_filename)

    # 3. Read file content to get size and save it
    try:
        content = await file.read()
        file_size = len(content)
        # Enforce file size limit of 5MB
        if file_size > 5 * 1024 * 1024:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Resume file size exceeds the 5MB limit."
            )
            
        with open(saved_file_path, "wb") as f:
            f.write(content)
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save resume file: {str(e)}"
        )

    # 4. Check if this is the first resume (make it active by default)
    try:
        existing_resumes = database.get_all_resumes(user_id)
        is_active = 1 if len(existing_resumes) == 0 else 0
        
        resume_payload = {
            "filename": filename,
            "filepath": f"/uploads/resumes/{disk_filename}",
            "file_size": file_size,
            "is_active": is_active,
            "user_id": user_id
        }
        
        resume_id = database.add_resume(resume_payload)
        saved_resume = database.get_resume_by_id(resume_id, user_id)
        if not saved_resume:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Saved resume details could not be retrieved"
            )
        jobs_cache.delete(f"resumes:{user_id}")  # Invalidate resumes list cache
        return saved_resume
    except Exception as e:
        # Clean up file if DB save failed
        if os.path.exists(saved_file_path):
            os.remove(saved_file_path)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save resume metadata: {str(e)}"
        )

@app.get("/api/resumes", response_model=List[models.ResumeDB])
def get_resumes(user_id: str = Depends(get_current_user)):
    """Retrieves all uploaded resumes. Cached per user for 30 seconds."""
    cache_key = f"resumes:{user_id}"
    cached = jobs_cache.get(cache_key)
    if cached is not None:
        return cached
    try:
        result = database.get_all_resumes(user_id)
        jobs_cache.set(cache_key, result)
        return result
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve resumes: {str(e)}"
        )

@app.post("/api/resumes/{resume_id}/active", response_model=models.ResumeDB)
def make_resume_active(resume_id: str, user_id: str = Depends(get_current_user)):
    """Sets a specific resume as active and deactivates all others."""
    try:
        resume = database.get_resume_by_id(resume_id, user_id)
        if not resume:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Resume with ID {resume_id} not found"
            )
        success = database.set_active_resume(resume_id, user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to update active resume in database"
            )
        jobs_cache.delete(f"resumes:{user_id}")  # Invalidate resumes list cache
        return database.get_resume_by_id(resume_id, user_id)
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update active resume status: {str(e)}"
        )

@app.delete("/api/resumes/{resume_id}")
def delete_resume_api(resume_id: str, user_id: str = Depends(get_current_user)):
    """Deletes a resume record from the database and its physical file from disk."""
    try:
        resume = database.get_resume_by_id(resume_id, user_id)
        if not resume:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Resume with ID {resume_id} not found"
            )

        # 1. Delete the physical file if it exists
        rel_filepath = resume.get("filepath")
        if rel_filepath:
            # Strip leading slash if present
            if rel_filepath.startswith("/"):
                rel_filepath = rel_filepath[1:]
            full_path = os.path.abspath(os.path.join(BASE_DIR, rel_filepath))
            if verify_path_safety(full_path, UPLOAD_DIR):
                if os.path.exists(full_path):
                    try:
                        os.remove(full_path)
                    except Exception:
                        pass  # Fail silently on disk removal, proceed to DB removal
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid or unauthorized resume file path"
                )
                    
        # 2. Delete the record from DB
        was_active = resume.get("is_active") == 1
        success = database.delete_resume(resume_id, user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to delete resume record from database"
            )
            
        jobs_cache.delete(f"resumes:{user_id}")  # Invalidate resumes list cache
        # 3. If the deleted resume was active, set the next most recent one as active
        if was_active:
            remaining = database.get_all_resumes(user_id)
            if remaining:
                # Set the first item (most recent upload) as active
                database.set_active_resume(remaining[0]["id"], user_id)
                
        return {"status": "success", "message": f"Resume {resume_id} deleted successfully"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete resume: {str(e)}"
        )

@app.post("/api/validate-key")
def validate_gemini_key():
    """
    Validates the server's Gemini API key by running a tiny test generation call.
    """
    api_key = get_gemini_key()
    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="API Key is missing on server."
        )
    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        
        # Determine best available model
        available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
        model_name = "gemini-1.5-flash"
        if available_models:
            model_name = available_models[0].replace('models/', '')
            
        model = genai.GenerativeModel(model_name)
        response = model.generate_content("Respond with exactly 'OK'")
        if response.text and "OK" in response.text:
            return {"valid": True, "message": "Server API Key is valid and active!"}
        else:
            return {"valid": True, "message": "Server API Key connected, but returned unexpected response."}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Server API Key error or connection error: {str(e)}"
        )

@app.get("/api/download-apk")
def download_apk():
    """Proxies the latest APK download from GitHub to avoid redirects."""
    url = "https://github.com/Harshvardhan210/MagicCounter/releases/latest/download/app-debug.apk"
    try:
        # Use streaming to avoid loading the whole APK into memory
        response = requests.get(url, stream=True, timeout=30)
        response.raise_for_status()
        
        headers = {
            "Content-Disposition": "attachment; filename=ZenJob.apk",
            "Content-Type": "application/vnd.android.package-archive"
        }
        
        return StreamingResponse(
            response.iter_content(chunk_size=1024 * 50), # 50KB chunks
            media_type="application/vnd.android.package-archive",
            headers=headers
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch APK from GitHub: {str(e)}"
        )

